"""
ZYN Sales Intelligence — Gerador de Relatórios Excel
Exporta análises em planilhas formatadas no padrão ZYN.
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR

# === Cores ZYN ===
NAVY = "223040"
GRAY = "8B9197"
GREEN = "2E7D4F"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F3F4"

HEADER_FONT = Font(name="Montserrat", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
ZEBRA_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
DATA_FONT = Font(name="Montserrat", size=9)
TITLE_FONT = Font(name="Montserrat", size=14, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Montserrat", size=11, color=GRAY)
THIN_BORDER = Border(
    bottom=Side(style="thin", color=GRAY),
)


def fmt_brl(value: float) -> str:
    """Formata valor em BRL."""
    if pd.isna(value) or value == 0:
        return "-"
    if abs(value) >= 1e9:
        return f"R$ {value/1e9:.2f}B"
    if abs(value) >= 1e6:
        return f"R$ {value/1e6:.1f}M"
    if abs(value) >= 1e3:
        return f"R$ {value/1e3:.0f}K"
    return f"R$ {value:,.0f}"


def apply_header_style(ws, row_num: int, max_col: int, fill=None):
    """Aplica estilo de cabeçalho a uma linha."""
    fill = fill or HEADER_FILL
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-ajusta largura das colunas."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def export_investor_profiles(profiles: pd.DataFrame, filename: str = None) -> Path:
    """Exporta perfis de investidores para Excel."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if filename is None:
        filename = f"ZYN_Investidores_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb = Workbook()

    # === Aba 1: Ranking Geral ===
    ws = wb.active
    ws.title = "Ranking Gestoras"

    # Título
    ws.merge_cells("A1:J1")
    ws["A1"] = "ZYN Sales Intelligence — Ranking de Gestoras"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Dados CVM (últimos 3 meses)"
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # Dados
    display_cols = {
        "gestora": "Gestora",
        "n_fundos": "Fundos",
        "pl_total": "PL Total",
        "vol_total": "Vol. RF Estruturada",
        "vol_NC": "Vol. NC",
        "vol_CRI": "Vol. CRI",
        "vol_CRA": "Vol. CRA",
        "vol_CPR-F": "Vol. CPR-F",
        "vol_DEBENTURE": "Vol. Debênture",
        "tipo_preferido": "Tipo Preferido",
        "ticket_medio": "Ticket Médio",
        "indexador_principal": "Indexador",
        "classe_predominante": "Classe",
    }
    available = {k: v for k, v in display_cols.items() if k in profiles.columns}

    # Header
    row = 4
    for col_idx, col_name in enumerate(available.values(), 1):
        ws.cell(row=row, column=col_idx, value=col_name)
    apply_header_style(ws, row, len(available))

    # Data
    money_cols = {"pl_total", "vol_total", "vol_NC", "vol_CRI", "vol_CRA", "vol_CPR-F", "vol_DEBENTURE", "ticket_medio"}
    for i, (_, profile) in enumerate(profiles.iterrows()):
        r = row + 1 + i
        for col_idx, col_key in enumerate(available.keys(), 1):
            val = profile.get(col_key, "")
            if col_key in money_cols and pd.notna(val) and val != 0:
                val = fmt_brl(float(val))
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL

    auto_width(ws)

    # === Aba 2: Por Tipo de Ativo ===
    for asset_type in ["NC", "CRI", "CRA", "CPR-F", "DEBENTURE"]:
        vol_col = f"vol_{asset_type}"
        if vol_col not in profiles.columns:
            continue
        subset = profiles[profiles[vol_col] > 0].sort_values(vol_col, ascending=False).head(50)
        if subset.empty:
            continue

        ws2 = wb.create_sheet(title=f"Top {asset_type}")
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"Top Compradores — {asset_type}"
        ws2["A1"].font = TITLE_FONT

        cols2 = {
            "gestora": "Gestora",
            "n_fundos": "Fundos",
            vol_col: f"Volume {asset_type}",
            f"n_ops_{asset_type}": "Nº Operações",
            "ticket_medio": "Ticket Médio",
            "indexador_principal": "Indexador",
            "pl_total": "PL Total",
        }
        avail2 = {k: v for k, v in cols2.items() if k in subset.columns}

        r = 3
        for col_idx, col_name in enumerate(avail2.values(), 1):
            ws2.cell(row=r, column=col_idx, value=col_name)
        apply_header_style(ws2, r, len(avail2))

        for i, (_, row_data) in enumerate(subset.iterrows()):
            for col_idx, col_key in enumerate(avail2.keys(), 1):
                val = row_data.get(col_key, "")
                if col_key in money_cols and pd.notna(val) and val != 0:
                    val = fmt_brl(float(val))
                cell = ws2.cell(row=r + 1 + i, column=col_idx, value=val)
                cell.font = DATA_FONT
                if i % 2 == 1:
                    cell.fill = ZEBRA_FILL

        auto_width(ws2)

    wb.save(filepath)
    print(f"📊 Relatório salvo: {filepath}")
    return filepath


def export_deal_matching(
    deal: dict,
    matching: pd.DataFrame,
    filename: str = None,
) -> Path:
    """Exporta matching de um deal específico para Excel."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    deal_name = deal.get("nome", "deal").replace(" ", "_").replace("/", "-")[:30]
    if filename is None:
        filename = f"ZYN_Match_{deal_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Matching"

    # Header do deal
    ws.merge_cells("A1:H1")
    ws["A1"] = f"ZYN Sales Intelligence — Matching para {deal.get('nome', 'N/A')}"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Tipo:"
    ws["B3"] = deal.get("tipo_raw", deal.get("tipo", ""))
    ws["C3"] = "Volume:"
    ws["D3"] = fmt_brl(deal.get("volume", 0))
    ws["E3"] = "Prazo:"
    ws["F3"] = f"{deal.get('prazo_anos', 'N/A')} anos" if deal.get("prazo_anos") else "N/A"
    ws["G3"] = "Indexador:"
    ws["H3"] = deal.get("indexador", "N/A")
    for cell in [ws["A3"], ws["C3"], ws["E3"], ws["G3"]]:
        cell.font = Font(name="Montserrat", size=9, bold=True, color=NAVY)

    # Ranking
    display_cols = {
        "gestora": "Gestora",
        "score_total": "Score",
        "n_fundos": "Fundos",
        "pl_total": "PL Total",
        "vol_total_rf": "Vol. RF Estruturada",
        "ticket_medio": "Ticket Médio",
        "tipo_preferido": "Tipo Preferido",
        "indexador_principal": "Indexador",
        "score_tipo": "S.Tipo",
        "score_volume": "S.Volume",
        "score_prazo": "S.Prazo",
    }
    avail = {k: v for k, v in display_cols.items() if k in matching.columns}

    r = 5
    for col_idx, col_name in enumerate(avail.values(), 1):
        ws.cell(row=r, column=col_idx, value=col_name)
    apply_header_style(ws, r, len(avail))

    for i, (_, row_data) in enumerate(matching.iterrows()):
        for col_idx, col_key in enumerate(avail.keys(), 1):
            val = row_data.get(col_key, "")
            if col_key in ("pl_total", "vol_total_rf", "ticket_medio") and pd.notna(val) and val != 0:
                val = fmt_brl(float(val))
            elif col_key.startswith("score_") and pd.notna(val):
                val = f"{float(val):.0%}"
            cell = ws.cell(row=r + 1 + i, column=col_idx, value=val)
            cell.font = DATA_FONT
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
            # Highlight score alto
            if col_key == "score_total" and isinstance(row_data.get(col_key), (int, float)):
                score_val = float(row_data[col_key])
                if score_val >= 0.7:
                    cell.font = Font(name="Montserrat", size=9, bold=True, color=GREEN)
                elif score_val >= 0.5:
                    cell.font = Font(name="Montserrat", size=9, color=NAVY)

    auto_width(ws)
    wb.save(filepath)
    print(f"📊 Matching salvo: {filepath}")
    return filepath
